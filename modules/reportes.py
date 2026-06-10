import customtkinter as ctk
from database.db import conectar
from datetime import datetime, timedelta
import openpyxl
import os
import json
import sqlite3
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


def get_idioma():
    try:
        config_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..",
            "config",
            "settings.json"
        )
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f).get("idioma", "Español")
    except Exception:
        return "Español"


def mostrar_alerta(titulo, mensaje, parent=None):
    ventana = ctk.CTkToplevel(parent) if parent else ctk.CTkToplevel()
    ventana.title(titulo)
    ventana.geometry("390x170")
    ventana.grab_set()
    ventana.resizable(False, False)

    ctk.CTkLabel(
        ventana,
        text=mensaje,
        text_color="#E8751A",
        font=ctk.CTkFont(size=13, weight="bold"),
        wraplength=350
    ).pack(pady=25)

    ctk.CTkButton(
        ventana,
        text="OK",
        fg_color="#E8751A",
        hover_color="#c45e0e",
        command=ventana.destroy
    ).pack()


def limpiar_contenido(content):
    for widget in content.winfo_children():
        if hasattr(widget, "_es_reporte"):
            widget.destroy()


def crear_canvas_grafica(fig, frame):
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="x", padx=15, pady=(0, 12))
    plt.close(fig)


def mostrar_reportes(frame):
    from config.translations import t
    idioma = get_idioma()

    content = ctk.CTkFrame(frame, fg_color="#0f0f0f")
    content.pack(fill="both", expand=True, padx=20, pady=20)

    ctk.CTkLabel(
        content,
        text=t("reportes", idioma),
        font=ctk.CTkFont(size=18, weight="bold"),
        text_color="#FFFFFF"
    ).pack(anchor="w", pady=(0, 15))

    botones_frame = ctk.CTkFrame(content, fg_color="transparent")
    botones_frame.pack(fill="x", pady=(0, 15))

    reportes = [
        ("📊 Resumen", lambda: mostrar_reporte_ventas(content)),
        ("🏆 Más vendidos", lambda: mostrar_productos_mas_vendidos(content)),
        ("💰 Ganancias", lambda: mostrar_ganancias(content)),
        ("📦 Stock bajo", lambda: mostrar_stock_bajo(content)),
        ("👥 Clientes", lambda: mostrar_clientes_frecuentes(content)),
        ("📚 Inventario", lambda: mostrar_valor_inventario(content)),
        ("⬇ Exportar", lambda: exportar_todo(content)),
    ]

    for texto, comando in reportes:
        ctk.CTkButton(
            botones_frame,
            text=texto,
            fg_color="#1a1a1a",
            hover_color="#2a2a2a",
            text_color="#FFFFFF",
            border_color="#E8751A",
            border_width=1,
            command=comando
        ).pack(side="left", padx=5)

    mostrar_reporte_ventas(content)


def obtener_stats_ventas():
    hoy = datetime.now().strftime("%Y-%m-%d")
    semana = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    mes = datetime.now().strftime("%Y-%m")

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*), COALESCE(SUM(total), 0) FROM ventas WHERE fecha LIKE ?", (f"{hoy}%",))
    v_hoy = cursor.fetchone()

    cursor.execute("SELECT COUNT(*), COALESCE(SUM(total), 0) FROM ventas WHERE fecha >= ?", (semana,))
    v_semana = cursor.fetchone()

    cursor.execute("SELECT COUNT(*), COALESCE(SUM(total), 0) FROM ventas WHERE fecha LIKE ?", (f"{mes}%",))
    v_mes = cursor.fetchone()

    cursor.execute("""
        SELECT DATE(fecha), COALESCE(SUM(total), 0)
        FROM ventas
        WHERE fecha >= ?
        GROUP BY DATE(fecha)
        ORDER BY DATE(fecha) ASC
    """, (semana,))
    grafica = cursor.fetchall()

    conn.close()
    return v_hoy, v_semana, v_mes, grafica


def mostrar_reporte_ventas(content):
    from config.translations import t
    idioma = get_idioma()
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    try:
        v_hoy, v_semana, v_mes, datos_grafica = obtener_stats_ventas()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    ctk.CTkLabel(
        frame,
        text=t("resumen_ventas", idioma),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    stats = [
        (t("hoy", idioma), v_hoy[0] or 0, v_hoy[1] or 0),
        (t("esta_semana", idioma), v_semana[0] or 0, v_semana[1] or 0),
        (t("este_mes", idioma), v_mes[0] or 0, v_mes[1] or 0),
    ]

    stats_frame = ctk.CTkFrame(frame, fg_color="transparent")
    stats_frame.pack(fill="x", padx=15, pady=(0, 12))

    for periodo, cantidad, total in stats:
        card = ctk.CTkFrame(stats_frame, fg_color="#222222", corner_radius=8)
        card.pack(side="left", expand=True, fill="x", padx=6)

        ctk.CTkLabel(card, text=periodo, text_color="#888888").pack(pady=(10, 2))
        ctk.CTkLabel(
            card,
            text=f"${total:.2f}",
            text_color="#E8751A",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack()
        ctk.CTkLabel(
            card,
            text=f"{cantidad} {t('ventas', idioma)}",
            text_color="#555555"
        ).pack(pady=(2, 10))

    if datos_grafica:
        dias = [d[0] for d in datos_grafica]
        totales = [d[1] or 0 for d in datos_grafica]

        fig, ax = plt.subplots(figsize=(8, 2.6))
        fig.patch.set_facecolor("#1a1a1a")
        ax.set_facecolor("#1a1a1a")
        ax.bar(dias, totales, color="#E8751A", alpha=0.85)
        ax.set_ylabel("Total $", color="#888888", fontsize=9)
        ax.tick_params(colors="#888888", labelsize=8)
        ax.spines["bottom"].set_color("#333333")
        ax.spines["left"].set_color("#333333")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        for label in ax.get_xticklabels():
            label.set_rotation(30)

        fig.tight_layout()
        crear_canvas_grafica(fig, frame)


def mostrar_productos_mas_vendidos(content):
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    ctk.CTkLabel(
        frame,
        text="🏆 Productos más vendidos",
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.nombre, COALESCE(SUM(vd.cantidad), 0) AS total_vendido,
                   COALESCE(SUM(vd.cantidad * vd.precio), 0) AS total_ingresos
            FROM venta_detalle vd
            LEFT JOIN productos p ON vd.producto_id = p.id
            GROUP BY vd.producto_id
            ORDER BY total_vendido DESC
            LIMIT 10
        """)
        productos = cursor.fetchall()
        conn.close()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    if not productos:
        ctk.CTkLabel(frame, text="Sin ventas registradas", text_color="#555555").pack(pady=30)
        return

    nombres = [(p[0] or "Producto eliminado")[:18] for p in productos]
    cantidades = [p[1] or 0 for p in productos]

    fig, ax = plt.subplots(figsize=(8, 2.8))
    fig.patch.set_facecolor("#1a1a1a")
    ax.set_facecolor("#1a1a1a")
    ax.barh(nombres, cantidades, color="#E8751A", alpha=0.85)
    ax.tick_params(colors="#888888", labelsize=8)
    ax.spines["bottom"].set_color("#333333")
    ax.spines["left"].set_color("#333333")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    crear_canvas_grafica(fig, frame)

    tabla = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    tabla.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    for i, p in enumerate(productos, start=1):
        fila = ctk.CTkFrame(tabla, fg_color="#222222", corner_radius=6)
        fila.pack(fill="x", pady=3)

        ctk.CTkLabel(fila, text=f"#{i}", width=40, text_color="#E8751A").pack(side="left", padx=8)
        ctk.CTkLabel(fila, text=p[0] or "Producto eliminado", width=250, anchor="w", text_color="#FFFFFF").pack(side="left")
        ctk.CTkLabel(fila, text=f"{p[1]} piezas", width=100, anchor="w", text_color="#888888").pack(side="left")
        ctk.CTkLabel(fila, text=f"${p[2]:.2f}", width=120, anchor="w", text_color="#AAAAAA").pack(side="left")


def mostrar_ganancias(content):
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    ctk.CTkLabel(
        frame,
        text="💰 Ganancias estimadas",
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                COALESCE(SUM(vd.cantidad * vd.precio), 0) AS ingresos,
                COALESCE(SUM(vd.cantidad * p.precio_costo), 0) AS costos,
                COALESCE(SUM(vd.cantidad * (vd.precio - p.precio_costo)), 0) AS ganancia
            FROM venta_detalle vd
            LEFT JOIN productos p ON vd.producto_id = p.id
        """)
        ingresos, costos, ganancia = cursor.fetchone()
        conn.close()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    stats_frame = ctk.CTkFrame(frame, fg_color="transparent")
    stats_frame.pack(fill="x", padx=15, pady=(0, 12))

    stats = [
        ("Ingresos", ingresos or 0, "#E8751A"),
        ("Costos", costos or 0, "#A32D2D"),
        ("Ganancia", ganancia or 0, "#3B6D11"),
    ]

    for titulo, valor, color in stats:
        card = ctk.CTkFrame(stats_frame, fg_color="#222222", corner_radius=8)
        card.pack(side="left", expand=True, fill="x", padx=6)

        ctk.CTkLabel(card, text=titulo, text_color="#888888").pack(pady=(10, 2))
        ctk.CTkLabel(
            card,
            text=f"${valor:.2f}",
            text_color=color,
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=(2, 12))


def mostrar_stock_bajo(content):
    from config.translations import t
    idioma = get_idioma()
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    ctk.CTkLabel(
        frame,
        text=t("stock_bajo", idioma),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT nombre, marca, cantidad, stock_minimo
            FROM productos
            WHERE cantidad <= stock_minimo
            ORDER BY cantidad ASC
        """)
        productos = cursor.fetchall()
        conn.close()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    if productos:
        nombres = [p[0][:15] for p in productos]
        cantidades = [p[2] or 0 for p in productos]
        minimos = [p[3] or 0 for p in productos]

        fig, ax = plt.subplots(figsize=(8, 2.6))
        fig.patch.set_facecolor("#1a1a1a")
        ax.set_facecolor("#1a1a1a")

        x = range(len(nombres))
        ax.bar(x, cantidades, color="#A32D2D", alpha=0.85, label="Stock")
        ax.bar(x, minimos, color="#555555", alpha=0.4, label="Mínimo")

        ax.set_xticks(list(x))
        ax.set_xticklabels(nombres, rotation=30, fontsize=8)
        ax.tick_params(colors="#888888", labelsize=8)
        ax.spines["bottom"].set_color("#333333")
        ax.spines["left"].set_color("#333333")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        fig.tight_layout()
        crear_canvas_grafica(fig, frame)

    tabla = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    tabla.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    if not productos:
        ctk.CTkLabel(tabla, text=t("sin_productos_reg", idioma), text_color="#555555").pack(pady=20)
        return

    for p in productos:
        fila = ctk.CTkFrame(tabla, fg_color="#222222", corner_radius=6)
        fila.pack(fill="x", pady=3)

        ctk.CTkLabel(fila, text=p[0], width=220, anchor="w", text_color="#FFFFFF").pack(side="left", padx=8)
        ctk.CTkLabel(fila, text=p[1] or "", width=130, anchor="w", text_color="#888888").pack(side="left")
        ctk.CTkLabel(fila, text=f"Stock: {p[2]}", width=100, anchor="w", text_color="#A32D2D").pack(side="left")
        ctk.CTkLabel(fila, text=f"Min: {p[3]}", width=80, anchor="w", text_color="#555555").pack(side="left")


def mostrar_clientes_frecuentes(content):
    from config.translations import t
    idioma = get_idioma()
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    ctk.CTkLabel(
        frame,
        text=t("clientes_frecuentes", idioma),
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT notas, COUNT(*) as total, COALESCE(SUM(total), 0)
            FROM ventas
            WHERE notas IS NOT NULL AND notas != ''
            GROUP BY notas
            ORDER BY total DESC
            LIMIT 10
        """)
        clientes = cursor.fetchall()
        conn.close()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    tabla = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    tabla.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    if not clientes:
        ctk.CTkLabel(
            tabla,
            text="No hay datos suficientes para clientes frecuentes.",
            text_color="#555555"
        ).pack(pady=20)
        return

    for i, c in enumerate(clientes, start=1):
        fila = ctk.CTkFrame(tabla, fg_color="#222222", corner_radius=6)
        fila.pack(fill="x", pady=3)

        ctk.CTkLabel(fila, text=f"#{i}", text_color="#E8751A", width=40).pack(side="left", padx=8)
        ctk.CTkLabel(fila, text=c[0], text_color="#FFFFFF", width=300, anchor="w").pack(side="left")
        ctk.CTkLabel(fila, text=f"{c[1]} ventas", text_color="#888888", width=100, anchor="w").pack(side="left")
        ctk.CTkLabel(fila, text=f"${c[2]:.2f}", text_color="#AAAAAA", width=120, anchor="w").pack(side="left")


def mostrar_valor_inventario(content):
    limpiar_contenido(content)

    frame = ctk.CTkFrame(content, fg_color="#1a1a1a", corner_radius=10)
    frame._es_reporte = True
    frame.pack(fill="both", expand=True, pady=(10, 0))

    ctk.CTkLabel(
        frame,
        text="📚 Valor del inventario",
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#E8751A"
    ).pack(anchor="w", padx=15, pady=(12, 8))

    try:
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                COUNT(*),
                COALESCE(SUM(cantidad), 0),
                COALESCE(SUM(cantidad * precio_costo), 0),
                COALESCE(SUM(cantidad * precio_menudeo), 0)
            FROM productos
        """)
        total_productos, total_piezas, valor_costo, valor_venta = cursor.fetchone()
        conn.close()
    except Exception as e:
        mostrar_alerta("Error BD", str(e), content)
        return

    stats_frame = ctk.CTkFrame(frame, fg_color="transparent")
    stats_frame.pack(fill="x", padx=15, pady=(0, 12))

    stats = [
        ("Productos", total_productos or 0),
        ("Piezas", total_piezas or 0),
        ("Valor costo", f"${valor_costo:.2f}"),
        ("Valor venta", f"${valor_venta:.2f}"),
    ]

    for titulo, valor in stats:
        card = ctk.CTkFrame(stats_frame, fg_color="#222222", corner_radius=8)
        card.pack(side="left", expand=True, fill="x", padx=6)

        ctk.CTkLabel(card, text=titulo, text_color="#888888").pack(pady=(10, 2))
        ctk.CTkLabel(
            card,
            text=str(valor),
            text_color="#E8751A",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=(2, 12))


def exportar_todo(parent=None):
    from config.translations import t
    idioma = get_idioma()

    try:
        conn = conectar()
        cursor = conn.cursor()

        wb = openpyxl.Workbook()

        ws_ventas = wb.active
        ws_ventas.title = "Ventas"
        ws_ventas.append(["ID", "Fecha", "Total", "Notas"])

        cursor.execute("SELECT id, fecha, total, notas FROM ventas ORDER BY id DESC")
        for v in cursor.fetchall():
            ws_ventas.append([v[0], v[1], v[2], v[3] or ""])

        ws_inv = wb.create_sheet("Inventario")
        ws_inv.append([
            "Nombre",
            "Marca",
            "SKU",
            "Cantidad",
            "Precio costo",
            "Precio menudeo",
            "Precio mayoreo",
            "Stock mínimo"
        ])

        cursor.execute("""
            SELECT nombre, marca, sku, cantidad, precio_costo,
                   precio_menudeo, precio_mayoreo, stock_minimo
            FROM productos
            ORDER BY nombre ASC
        """)
        for p in cursor.fetchall():
            ws_inv.append(list(p))

        ws_clientes = wb.create_sheet("Clientes")
        ws_clientes.append(["Nombre", "Teléfono", "Email", "Dirección"])

        cursor.execute("SELECT nombre, telefono, email, direccion FROM clientes ORDER BY nombre")
        for c in cursor.fetchall():
            ws_clientes.append(list(c))

        ws_prov = wb.create_sheet("Proveedores")
        ws_prov.append(["Nombre", "Teléfono", "Email", "Dirección"])

        cursor.execute("SELECT nombre, telefono, email, direccion FROM proveedores ORDER BY nombre")
        for p in cursor.fetchall():
            ws_prov.append(list(p))

        ws_apartados = wb.create_sheet("Apartados")
        ws_apartados.append(["Cliente", "Producto", "Cantidad", "Fecha", "Entrega", "Estado"])

        cursor.execute("""
            SELECT a.cliente, p.nombre, a.cantidad, a.fecha, a.fecha_entrega, a.estado
            FROM apartados a
            LEFT JOIN productos p ON a.producto_id = p.id
            ORDER BY a.id DESC
        """)
        for a in cursor.fetchall():
            ws_apartados.append(list(a))

        conn.close()

        hoy = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        nombre = f"reporte_completo_{hoy}.xlsx"
        ruta = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..",
            "exports",
            nombre
        )

        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        wb.save(ruta)

        mostrar_alerta(
            "Exportación completa",
            f"Archivo guardado en:\nexports/{nombre}",
            parent
        )

    except PermissionError:
        mostrar_alerta(
            "Error",
            "No se pudo guardar el Excel.\nCierra el archivo si está abierto.",
            parent
        )

    except Exception as e:
        mostrar_alerta("Error", f"No se pudo exportar:\n{e}", parent)